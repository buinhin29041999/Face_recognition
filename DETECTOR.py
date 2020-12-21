import numpy as np
from cv2 import cv2
import sqlite3

import openpyxl

face_cascade = cv2.CascadeClassifier('lib/haarcascade_frontalface_default.xml')
cap = cv2.VideoCapture(0)
rec = cv2.face.LBPHFaceRecognizer_create()
rec.read("trainingdata.yml")
id = 0

workbook = openpyxl.load_workbook('Diemdanh.xlsx')
ws = workbook.worksheets[0]

# set text style
fontface = cv2.FONT_HERSHEY_COMPLEX_SMALL
fontscale = 1
fontcolor = (255, 23, 252)


def getProfile(id):
    conn = sqlite3.connect("FaceBase.db")
    cmd = "SELECT * FROM People WHERE ID="+str(id)
    cursor = conn.execute(cmd)
    profile = None
    for row in cursor:
        profile = row
    conn.close()
    return profile

while 1:
    ret, img = cap.read()
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.5, 5)

    for (x, y, w, h) in faces:
        cv2.rectangle(img, (x, y), (x+w, y+h), (255, 255, 0), 2)
        id, conf = rec.predict(gray[y:y+h, x:x+w])
        profile = getProfile(id)
        # Start from the first cell. Rows and columns are zero indexed.
        hang = id
        cot = 1
        # set text to window
        if(profile != None):
            cv2.putText(
                img, "Name: " + str(profile[1]), (x+w+10, y+10), fontface, fontscale, fontcolor, 2)
            cv2.putText(
                img, "IDSV: " + str(profile[2]), (x+w+10, y+30), fontface, fontscale, fontcolor, 2)
            cv2.putText(
                img, "Gender: " + str(profile[3]), (x+w+10, y+50), fontface, fontscale, fontcolor, 2)
            # Iterate over the data and write it out row by row.
            if(ws.cell(row=hang, column=1).value == None):
                ws.cell(row=hang, column=cot).value = str(profile[2])
                ws.cell(row=hang, column=cot+1).value = str(profile[1])
                ws.cell(row=hang, column=cot+2).value = str(profile[3])
                workbook.save('Diemdanh.xlsx')

    cv2.imshow('img', img)
    if cv2.waitKey(1) == ord('q'):
        break


cap.release()
cv2.destroyAllWindows()
